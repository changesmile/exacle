import openpyxl

workbook = openpyxl.load_workbook('test_log.xlsx')
sheetname = workbook.sheetnames
worksheet = workbook[sheetname[0]]
cInvCode = [item.value for item in list(worksheet.columns)[0]]
# 存入物料码
proDesc = [item.value for item in list(worksheet.columns)[15]]
# 获取工序说明
cInvCode_flag = cInvCode[4]  # 取出第一个物料码
proDesc_flag = proDesc[4]  # 取出第一个工序说明
proDesc_all = []
flag_all = []
flag = 1  # 记录下标
str_flag = proDesc_flag  # 创立一个符串，将第一个工序说明导入
for i, j in zip(cInvCode[5:], proDesc[5:]):
    if i == cInvCode_flag:
        str_flag = str_flag + "." + j
    elif i != cInvCode_flag:
        proDesc_all.append(str_flag)
        flag_all.append(flag)
        # print(str_flag)
        str_flag = j  # 重新存储第一个工序说明
        cInvCode_flag = i
    flag = flag + 1
# print(str_flag)
proDesc_all.append(str_flag)
flag_all.append(flag)
print(proDesc_all)
print("----------")
print(flag_all)

# cell_2_3 = worksheet.cell(row=2, column=3).value
# print('第2行第3列值', cell_2_3)


wb1 = workbook.active  # 激活sheet
for i, j in zip(proDesc_all, flag_all):
    wb1.cell(j+4, 41, i)
workbook.save("test.xlsx")  # 保存