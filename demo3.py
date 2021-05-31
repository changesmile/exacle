from treelib import Node, Tree
import openpyxl

pro_tree = Tree()
pro_tree.create_node(tag="Root", identifier="root")
log_wk = openpyxl.load_workbook('test_log.xlsx')
log_name = log_wk.sheetnames
log_ws = log_wk[log_name[0]]

pro_wk = openpyxl.load_workbook('test_project.xlsx')
pro_name = pro_wk.sheetnames
pro_ws = pro_wk[pro_name[0]]
flag = 0


def bijiao(str):
    for j in log_ws.values:
        if str == j[1]:
            return j[3]
            break
    return "None"


for i in pro_ws.values:
    if flag == 0:
        flag = 1
        continue
    log_time = bijiao(i[1])
    # print(log_time)
    if not pro_tree.contains(nid=i[3]):
        pro_tree.create_node(tag=i[3], identifier=i[3], parent="root")
        # if not pro_tree.contains(nid=i[5]):
            # log_time = bijiao(i[1])
    pro_tree.create_node(tag=f"{i[5]}--{log_time}", identifier=i[5], parent=i[3],)

pro_tree.show("root")